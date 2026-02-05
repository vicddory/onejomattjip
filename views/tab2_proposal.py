# -*- coding: utf-8 -*-
"""
================================================================================
📁 views/tab2_proposal.py - 커피 무역 제안서 생성기
================================================================================
[리팩토링 v10] UI 레이아웃 변경
- 단가 표시 위치 이동: 왼쪽 -> 오른쪽
- 수입 물량 섹션(오른쪽)을 2x2 그리드(단가, 중량 / USD총액, KRW총액)로 재구성
================================================================================
"""


import streamlit as st
import os
from datetime import datetime
from io import BytesIO


# PDF 라이브러리
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


# Excel 라이브러리
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# 경로 설정
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


# API 키 및 유틸리티 가져오기
try:
    from config import OPENAI_API_KEY
    from utils import get_exchange_rate_with_status, get_country_weather
except ImportError:
    # 파일이 없거나 에러날 경우를 대비한 더미 데이터
    OPENAI_API_KEY = None
    def get_exchange_rate_with_status(): return 1450.0, "API 미연동 (기본값)"
    def get_country_weather(city): return {'temp': 20, 'desc_ko': '맑음', 'desc_en': 'Clear'}




# ===========================================
# 폰트 설정 (PDF용 - 한글 깨짐 방지)
# ===========================================
KOREAN_FONT = 'Helvetica'
USE_KOREAN_FONT = False


def register_korean_font():
    global KOREAN_FONT, USE_KOREAN_FONT
    font_candidates = [
        '/usr/share/fonts/truetype/nanum/NanumGothic.ttf',
        'C:/Windows/Fonts/malgun.ttf',
        'C:/Windows/Fonts/Gulim.ttc'
    ]
    for path in font_candidates:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont('KoreanFont', path))
                KOREAN_FONT = 'KoreanFont'
                USE_KOREAN_FONT = True
                return
            except:
                continue


register_korean_font()




# ===========================================
# 커피 품종 데이터
# ===========================================
def get_coffee_varieties():
    return {
        "에티오피아": {
            "port": "Djibouti", "port_en": "Djibouti Port", "country_en": "Ethiopia",
            "varieties": {
                "Gesha (게이샤)": {"price": 12.5, "desc": "자스민 향과 독보적인 산미", "desc_en": "Jasmine aroma and distinctive acidity"},
                "Yirgacheffe (예가체프)": {"price": 6.8, "desc": "꽃향기와 밝은 산미", "desc_en": "Floral aroma and bright acidity"},
                "Sidamo (시다모)": {"price": 5.8, "desc": "풍부한 과일 향과 바디감", "desc_en": "Fruity aroma with full body"}
            }
        },
        "케냐": {
            "port": "Mombasa", "port_en": "Mombasa Port", "country_en": "Kenya",
            "varieties": {
                "SL28": {"price": 8.5, "desc": "와인 풍미와 강렬한 산미", "desc_en": "Wine-like flavor with intense acidity"},
                "SL34": {"price": 8.2, "desc": "묵직한 바디감과 깊은 풍미", "desc_en": "Full body with deep flavor"},
                "Ruiru 11": {"price": 5.5, "desc": "깔끔한 맛과 적절한 산미", "desc_en": "Balanced acidity with clean taste"}
            }
        },
        "콜롬비아": {
            "port": "Buenaventura", "port_en": "Buenaventura Port", "country_en": "Colombia",
            "varieties": {
                "Typica (티피카)": {"price": 6.5, "desc": "깔끔한 향미와 단맛", "desc_en": "Clean flavor with sweet finish"},
                "Caturra (카투라)": {"price": 5.2, "desc": "풍부한 산미와 중간 바디감", "desc_en": "Rich acidity with medium body"},
                "Castillo (카스티요)": {"price": 5.0, "desc": "부드럽고 베리류 향미", "desc_en": "Smooth with berry notes"}
            }
        },
        "과테말라": {
            "port": "Puerto Barrios", "port_en": "Puerto Barrios Port", "country_en": "Guatemala",
            "varieties": {
                "Pacamara (파카마라)": {"price": 7.2, "desc": "복합적인 꽃향기와 묵직한 바디", "desc_en": "Complex floral aroma and full body"},
                "Antigua (안티구아)": {"price": 5.4, "desc": "스모키한 향과 초콜릿 풍미", "desc_en": "Smoky aroma with chocolate flavor"},
                "Bourbon (버번)": {"price": 5.2, "desc": "고소함과 산미의 조화", "desc_en": "Nutty sweetness with smooth acidity"}
            }
        },
        "브라질": {
            "port": "Santos", "port_en": "Santos Port", "country_en": "Brazil",
            "varieties": {
                "Bourbon (버번)": {"price": 5.2, "desc": "뛰어난 단맛과 밸런스", "desc_en": "Excellent sweetness and balance"},
                "Catuai (카투아이)": {"price": 4.5, "desc": "가벼운 바디감과 깔끔함", "desc_en": "Light body and clean taste"},
                "Mundo Novo (문도노보)": {"price": 4.2, "desc": "생산성 좋고 밸런스 잡힘", "desc_en": "Productive and well-balanced"}
            }
        },
        "인도네시아": {
            "port": "Jakarta", "port_en": "Jakarta Port", "country_en": "Indonesia",
            "varieties": {
                "Mandheling (만델링)": {"price": 5.5, "desc": "흙내음과 초콜릿, 묵직함", "desc_en": "Earthy, chocolate notes with heavy body"},
                "Lintong (린통)": {"price": 5.2, "desc": "허브 향과 묵직한 질감", "desc_en": "Herbal aroma with heavy texture"},
                "Gayo (가요)": {"price": 4.9, "desc": "산미와 단맛의 좋은 균형", "desc_en": "Balanced acidity and sweetness"}
            }
        },
        "베트남": {
            "port": "Ho Chi Minh", "port_en": "Ho Chi Minh Port", "country_en": "Vietnam",
            "varieties": {
                "Excelsa (엑셀사)": {"price": 4.2, "desc": "독특한 과일 향과 타르트 산미", "desc_en": "Unique fruity aroma with tart acidity"},
                "Catimor (카티모르)": {"price": 3.8, "desc": "산미와 쓴맛의 밸런스", "desc_en": "Balanced acidity and bitterness"},
                "Robusta (로부스타)": {"price": 3.2, "desc": "강한 바디감과 구수한 맛", "desc_en": "Strong body with savory taste"}
            }
        },
        "코스타리카": {
            "port": "Limon", "port_en": "Limon Port", "country_en": "Costa Rica",
            "varieties": {
                "Villa Sarchi (빌라 사치)": {"price": 7.5, "desc": "우아한 산미와 꽃향기", "desc_en": "Elegant acidity with floral notes"},
                "Caturra (카투라)": {"price": 5.9, "desc": "밝은 산미와 깨끗한 맛", "desc_en": "Bright acidity with clean finish"},
                "Venecia (베네치아)": {"price": 6.2, "desc": "깊은 단맛과 바디감", "desc_en": "Deep sweetness and full body"}
            }
        },
        "페루": {
            "port": "Callao", "port_en": "Callao Port", "country_en": "Peru",
            "varieties": {
                "Typica (티피카)": {"price": 5.1, "desc": "은은한 단맛과 깔끔함", "desc_en": "Subtle sweetness with clean finish"},
                "Bourbon (버번)": {"price": 4.8, "desc": "깊은 풍미와 밸런스", "desc_en": "Deep flavor with excellent balance"},
                "Pache (파체)": {"price": 4.5, "desc": "부드럽고 편안한 맛", "desc_en": "Smooth and mild flavor"}
            }
        },
        "온두라스": {
            "port": "Puerto Cortes", "port_en": "Puerto Cortes Port", "country_en": "Honduras",
            "varieties": {
                "Parainema (파라이네마)": {"price": 4.8, "desc": "열대 과일 향과 부드러움", "desc_en": "Tropical fruit aroma and smooth"},
                "Lempira (렘피라)": {"price": 4.1, "desc": "카라멜 단맛과 견과류", "desc_en": "Caramel sweetness with nutty flavor"},
                "Ihcatefe (이카페)": {"price": 3.9, "desc": "밝은 산미와 청량함", "desc_en": "Bright acidity with refreshing finish"}
            }
        }
    }




# ===========================================
# AI 전문가 분석 함수 (구체적 판단 로직 추가)
# ===========================================
def get_ai_advice(context_data, lang_code):
    """
    OpenAI API를 사용하여 무역 제안에 대한 구체적인 조언을 생성합니다.
    """
    if not OPENAI_API_KEY:
        return "⚠️ API KEY ERROR: .env 파일에 OPENAI_API_KEY가 없습니다. 키를 확인해주세요."


    try:
        from openai import OpenAI
        client = OpenAI(api_key=OPENAI_API_KEY)
       
        # 언어 설정 (강제성 부여)
        if lang_code == 'ko':
            system_instruction = "당신은 세계적인 커피 무역 전문가입니다. 한국어로 답변하세요."
            output_format = "다음 형식으로 답변해: 1. 시장성 분석, 2. 리스크 요인, 3. 최종 매수 추천 여부(강력 추천/보류/비추천)"
        else:
            system_instruction = "You are a world-class coffee trade expert. Respond ONLY in English."
            output_format = "Answer in this format: 1. Marketability Analysis, 2. Risk Factors, 3. Final Recommendation (Strong Buy/Hold/Don't Buy)"


        user_prompt = f"""
        Analyze this coffee import deal specifically:
        - Origin: {context_data['country_en']}
        - Variety: {context_data['variety_en']}
        - Price: ${context_data['unit_price']}/kg
        - Exchange Rate: {context_data['exchange_rate']} KRW/USD
        - Local Weather: {context_data['weather_en']}
       
        {output_format}
        Provide a sharp, professional business judgment in 3-4 sentences.
        """
       
        response = client.chat.completions.create(
            model="gpt-4o-mini", # 혹은 gpt-3.5-turbo
            messages=[
                {"role": "system", "content": system_instruction},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.7,
            max_tokens=300
        )
        return response.choices[0].message.content.strip()
       
    except Exception as e:
        return f"❌ AI 분석 실패: {str(e)} (API 키나 인터넷 연결을 확인하세요)"




# ==========================================
# PDF 생성 함수
# ==========================================
def create_pdf_proposal(data, lang='ko'):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
   
    styles = getSampleStyleSheet()
    font_name = KOREAN_FONT if USE_KOREAN_FONT else 'Helvetica'
    font_name_bold = KOREAN_FONT if USE_KOREAN_FONT else 'Helvetica-Bold'
   
    # 스타일 정의
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontName=font_name_bold, fontSize=24, textColor=colors.HexColor('#1F4788'), spaceAfter=20)
    h1_style = ParagraphStyle('H1', parent=styles['Heading1'], fontName=font_name_bold, fontSize=16, textColor=colors.HexColor('#2E5C8A'), spaceAfter=12)
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=font_name, fontSize=11, leading=16)
   
    story = []
    is_ko = (lang == 'ko')
   
    # 언어에 따른 텍스트 설정
    txt = {
        'title': "수입 의사결정 제안서" if is_ko else "Coffee Import Proposal",
        'date': f"제안일자: {data['date']}" if is_ko else f"Date: {data['date']}",
        's1': "1. 수입 개요" if is_ko else "1. Import Overview",
        's2': "2. 비용 및 규모" if is_ko else "2. Cost & Volume",
        's3': "3. 종합 의견 (AI Analysis)" if is_ko else "3. Recommendations",
        'footer': "본 제안서는 Coffee Trade Dashboard에서 생성되었습니다." if is_ko else "Generated by Coffee Trade Dashboard."
    }


    # 본문 작성
    story.append(Paragraph(txt['title'], title_style))
    story.append(Paragraph(txt['date'], ParagraphStyle('Date', parent=normal_style, alignment=TA_CENTER, textColor=colors.grey)))
    story.append(Spacer(1, 0.5*cm))
    story.append(Spacer(1, 1*cm))
   
    # 섹션 1
    story.append(Paragraph(txt['s1'], h1_style))
    country_val = f"{data['country']} ({data['port']}항)" if is_ko else f"{data['country_en']} ({data['port_en']})"
    variety_val = data['variety'] if is_ko else data['variety_en']
   
    tbl_data = [
        ["수입 대상국" if is_ko else "Origin Country", country_val],
        ["선택 품종" if is_ko else "Coffee Variety", variety_val],
        ["적용 환율" if is_ko else "Exchange Rate", f"{data['exchange_rate']} KRW/USD"]
    ]
   
    t = Table(tbl_data, colWidths=[4.5*cm, 12.5*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#E7F0F9')),
        ('FONTNAME', (0,0), (-1,-1), font_name),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 1*cm))
   
    # 섹션 2
    story.append(Paragraph(txt['s2'], h1_style))
    story.append(Paragraph(f"• {'단가' if is_ko else 'Unit Price'}: ${data['unit_price']}/kg", normal_style))
    story.append(Paragraph(f"• {'규모' if is_ko else 'Volume'}: {data['quantity_ton']} ton", normal_style))
    story.append(Paragraph(f"• {'총액' if is_ko else 'Total'}: ${data['total_usd']} (≈ {data['total_krw']} KRW)", normal_style))
    story.append(Spacer(1, 1*cm))
   
    # 섹션 3
    story.append(Paragraph(txt['s3'], h1_style))
   
    # AI 의견을 1, 2, 3번으로 분리
    ai_text = data['ai_opinion']
    paragraphs = []
   
    # 1., 2., 3. 으로 분리
    if '1. ' in ai_text and '2. ' in ai_text:
        parts = ai_text.split('2. ')
        part1 = parts[0].replace('1. ', '').strip()
       
        if '3. ' in parts[1]:
            sub_parts = parts[1].split('3. ')
            part2 = sub_parts[0].strip()
            part3 = sub_parts[1].strip()
           
            paragraphs = [
                f"<b>1.</b> {part1}",
                f"<b>2.</b> {part2}",
                f"<b>3.</b> {part3}"
            ]
        else:
            paragraphs = [
                f"<b>1.</b> {part1}",
                f"<b>2.</b> {parts[1].strip()}"
            ]
    else:
        paragraphs = [ai_text]
   
    # 각 문단을 개별 Paragraph로 추가
    for para_text in paragraphs:
        story.append(Paragraph(para_text, normal_style))
        story.append(Spacer(1, 0.3*cm))
   
    story.append(Spacer(1, 1.5*cm))
    story.append(Paragraph(txt['footer'], ParagraphStyle('Footer', parent=normal_style, alignment=TA_CENTER, fontSize=9, textColor=colors.grey)))


    doc.build(story)
    buffer.seek(0)
    return buffer




# ==========================================
# Excel 생성 함수
# ==========================================
def create_excel_proposal(data, lang='ko'):
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Proposal"
   
    is_ko = (lang == 'ko')
   
    # 스타일
    title_font = Font(name='맑은 고딕' if is_ko else 'Calibri', size=20, bold=True, color='1F4788')
    section_header_font = Font(name='맑은 고딕' if is_ko else 'Calibri', size=14, bold=True, color='1F4788')
    table_header_fill = PatternFill(start_color='E7F0F9', end_color='E7F0F9', fill_type='solid')
    red_bold_font = Font(name='맑은 고딕' if is_ko else 'Calibri', bold=True, color='C00000')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
   
    # 1. Title
    ws.merge_cells('A1:B1')
    ws['A1'] = "수입 의사결정 제안서" if is_ko else "Coffee Import Proposal"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
   
    # Date
    ws.merge_cells('A3:B3')
    ws['A3'] = f"Date: {data['date']}"
    ws['A3'].alignment = Alignment(horizontal='center')
   
    # 2. Import Overview
    row = 5
    ws[f'A{row}'] = "1. 수입 개요" if is_ko else "1. Import Overview"
    ws[f'A{row}'].font = section_header_font
    row += 1
   
    char_val = data.get('desc') if is_ko else data.get('desc_en')
    if not char_val: char_val = "-"
   
    labels_s1 = [
        ("수입 대상국" if is_ko else "Origin Country", f"{data['country']} ({data['port']}항)" if is_ko else f"{data['country_en']} ({data['port_en']})"),
        ("커피 품종" if is_ko else "Coffee Variety", data['variety'] if is_ko else data['variety_en']),
        ("특징" if is_ko else "Characteristics", char_val),
        ("적용 환율" if is_ko else "Exchange Rate", f"{data['exchange_rate']} KRW/USD")
    ]
   
    for label, value in labels_s1:
        cell_a = ws[f'A{row}']
        cell_a.value = label
        cell_a.fill = table_header_fill
        cell_a.border = thin_border
       
        cell_b = ws[f'B{row}']
        cell_b.value = value
        cell_b.border = thin_border
        row += 1
       
    row += 1
   
    # 3. Cost & Volume
    ws[f'A{row}'] = "2. 비용 및 규모" if is_ko else "2. Cost & Volume"
    ws[f'A{row}'].font = section_header_font
    row += 1
   
    labels_s2 = [
        ("• 단가" if is_ko else "• Unit Price", f"${data['unit_price']}/kg"),
        ("• 수입 물량" if is_ko else "• Import Volume", f"{data['quantity_ton']} ton"),
    ]
   
    for label, value in labels_s2:
        ws[f'A{row}'] = label
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'] = value
        ws[f'B{row}'].border = thin_border
        row += 1
       
    ws[f'A{row}'] = "• 예상 총액 (FOB)" if is_ko else "• Estimated Total (FOB)"
    ws[f'A{row}'].border = thin_border
   
    total_str = f"${data['total_usd']} ({data['total_krw']} KRW)"
    ws[f'B{row}'] = total_str
    ws[f'B{row}'].font = red_bold_font
    ws[f'B{row}'].border = thin_border
   
    row += 2
   
    # 4. Recommendations
    ws[f'A{row}'] = "3. 종합 의견" if is_ko else "3. Recommendations"
    ws[f'A{row}'].font = section_header_font
    row += 1
   
    ws.merge_cells(f'A{row}:B{row+2}')
    cell_advice = ws[f'A{row}']
    cell_advice.value = data['ai_opinion']
    cell_advice.alignment = Alignment(wrap_text=True, vertical='top')
   
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
   
    wb.save(buffer)
    buffer.seek(0)
    return buffer




# ===========================================
# 메인 show() 함수
# ===========================================
def show():
    """무역 제안서 생성기를 렌더링합니다."""
   
    data = get_coffee_varieties()
   
    # 세션 상태 초기화
    if 'exchange_source' not in st.session_state:
        st.session_state['exchange_source'] = 'manual'
    if 'api_rate' not in st.session_state:
        st.session_state['api_rate'] = None
    if 'proposal_manual_rate' not in st.session_state:
        st.session_state['proposal_manual_rate'] = 1450.0
   
    # 타이틀
    st.markdown("<h1 style='text-align: center; color:#6F4E37;'>국가별 원두 및 수입 제안서</h1>", unsafe_allow_html=True)
    st.markdown(" ")
    st.markdown(" ")    


    # ===========================================
    # 3개 컬럼 레이아웃 (스타일 업데이트: tab3_cost_calculator 스타일 적용)
    # ===========================================
    BOX_HEIGHT = 120
    col1, col2, col3 = st.columns([1, 1, 1], gap="medium")
   
    # 1. 오늘의 환율 (Col 1)
    with col1:
        with st.container(border=True, height=BOX_HEIGHT):
            st.markdown("##### 오늘의 환율")
           
            # 값 계산
            if st.session_state['exchange_source'] == 'api' and st.session_state['api_rate']:
                exchange_rate = st.session_state['api_rate']
            else:
                exchange_rate = st.session_state['proposal_manual_rate']
           
            # 레이아웃: 값(왼쪽) + 버튼(오른쪽 하단)
            c_val, c_btn = st.columns([3, 1], vertical_alignment="bottom")
            with c_val:
                st.metric(label="USD 기준", value=f"{exchange_rate:,.2f} 원", label_visibility="collapsed")
            with c_btn:
                if st.button("갱신 →", use_container_width=True, key="proposal_rate_btn"):
                    with st.spinner("환율 정보를 가져오는 중..."):
                        rate, msg = get_exchange_rate_with_status()
                        if rate:
                            st.session_state['api_rate'] = rate
                            st.session_state['exchange_source'] = 'api'
                            st.success(msg)
                            st.rerun()
                        else:
                            st.error(msg)
   
    # 2. 환율 수동 설정 (Col 2)
    with col2:
        with st.container(border=True, height=BOX_HEIGHT):
            st.markdown("##### 환율 수동 설정")
           
            # 레이아웃: 입력(왼쪽) + 버튼(오른쪽 하단)
            c_input, c_btn = st.columns([3, 1], vertical_alignment="bottom")
            with c_input:
                manual_rate = st.number_input(
                    "수동 환율 (KRW)",
                    min_value=100.0, max_value=10000.0,
                    value=st.session_state['proposal_manual_rate'],
                    step=10.0, format="%.2f",
                    key="manual_rate_input",
                    label_visibility="collapsed"
                )
            with c_btn:
                if st.button("적용 →", use_container_width=True, key="apply_manual_rate"):
                    st.session_state['proposal_manual_rate'] = manual_rate
                    st.session_state['exchange_source'] = 'manual'
                    st.rerun()


    # 3. 국가 선택 (Col 3) - 스타일 통일
    with col3:
        with st.container(border=True, height=BOX_HEIGHT):
            st.markdown("##### 국가 선택")
           
            # 상단 헤더와의 간격 및 입력창 배치를 위해 빈 공간 추가 or 하단 정렬 효과
            # Col1, Col2의 내부 컬럼과 높이를 맞추기 위해 간단히 selectbox 배치
            sorted_countries = sorted(list(data.keys()))
            country_options = [c for c in sorted_countries]
           
            # label_visibility="collapsed"로 통일감 부여
            selected_display = st.selectbox(
                "산지 국가",
                country_options,
                key="proposal_country_display",
                label_visibility="collapsed"
            )
            selected_country = selected_display
   
    # --------------------------------------------------------------------------
    # 적용된 환율 표시 (선택사항: tab3처럼 아래에 띠로 보여줄 수도 있음. 여기서는 간단히 Divider로 처리)
    st.write("") # 간격


    # [수정됨] 데이터 결정 로직 (이 값이 최종적으로 계산에 쓰임)
    if st.session_state['exchange_source'] == 'api' and st.session_state['api_rate']:
        final_applied_rate = st.session_state['api_rate']
    else:
        # [KeyError 수정 유지] Tab2에서는 proposal_manual_rate를 사용해야 합니다.
        final_applied_rate = st.session_state['proposal_manual_rate']    


    st.markdown(f"""
        <div style='
            background-color: #E0F2F1;
            border: 1px solid #B2DFDB;
            border-radius: 8px;
            padding: 12px;
            text-align: center;
            margin-top: 5px;
        '>
            <span style='color: #424242; font-size: 15px; font-weight: 500;'>현재 적용 환율:</span>
            <span style='color: #00695C; font-size: 18px; font-weight: 800; margin-left: 8px;'>
                {final_applied_rate:,.2f} 원/USD
            </span>
        </div>
    """, unsafe_allow_html=True)
   
    st.divider()


    country_info = data[selected_country]
    weather_data = get_country_weather(country_info['port'])
       
    # 산지 정보 표시
    st.subheader(f"산지 정보 - {selected_country}")
    info_col1, info_col2, info_col3 = st.columns(3)
    with info_col1: st.metric("항구", country_info['port'])
    with info_col2: st.metric("현지 기온", f"{weather_data['temp']}°C")
    with info_col3: st.metric("날씨", weather_data['desc_ko'])
   
    st.divider()
   
    # 품종 및 물량
    col_variety, col_quantity = st.columns(2)
    with col_variety:
        st.markdown("### 1. 품종 선택")
        selected_v = st.radio("커피 품종", list(country_info['varieties'].keys()), key="proposal_variety", label_visibility="collapsed")
        v_info = country_info['varieties'][selected_v]
        st.success(f"특징: {v_info['desc']}")
        # [수정] 단가 정보 삭제 (오른쪽으로 이동)
   
    with col_quantity:
        st.markdown("### 2. 수입 물량")
        qty = st.number_input("수입 물량 (Ton)", min_value=1.0, max_value=100.0, value=10.0, step=1.0, key="proposal_qty")
       
        price = v_info['price']
        total_usd = qty * 1000 * price
        total_krw = total_usd * exchange_rate
       
        # [수정] 2x2 그리드 레이아웃 적용
        # Row 1: 단가 | 총 중량
        q_row1_col1, q_row1_col2 = st.columns(2)
        with q_row1_col1:
            st.metric("단가", f"${price:.2f}/kg")
        with q_row1_col2:
            st.metric("총 중량", f"{qty * 1000:,.0f} kg")


        # Row 2: 총액 (USD) | 총액 (KRW)
        q_row2_col1, q_row2_col2 = st.columns(2)
        with q_row2_col1:
            st.metric("총액 (USD)", f"${total_usd:,.2f}")
        with q_row2_col2:
            st.metric("총액 (KRW)", f"{int(total_krw):,} 원")
   
    st.divider()
   
    # ===========================================
    # AI 제안서 생성
    # ===========================================
    st.markdown("### AI 제안서 생성")
   
    lang_choice = st.radio("언어 선택: ", ["한국어 (Korean)", "English"], horizontal=True, key="proposal_lang")
    lang_code = 'ko' if "한국어" in lang_choice else 'en'
   
    if 'generated_proposal' not in st.session_state:
        st.session_state['generated_proposal'] = None
        st.session_state['generated_lang'] = 'ko'
   
    if st.button("AI 제안서 생성하기", use_container_width=True, key="generate_proposal_btn"):
        with st.spinner("AI가 데이터를 분석하여 제안서를 작성 중입니다..."):
            weather_str = f"{weather_data['temp']}°C, {weather_data['desc_ko'] if lang_code == 'ko' else weather_data.get('desc_en', 'Clear')}"
           
            # 기본 데이터 구성
            prop_data = {
                'date': datetime.now().strftime('%Y-%m-%d'),
                'country': selected_country,
                'country_en': country_info['country_en'],
                'port': country_info['port'],
                'port_en': country_info['port_en'],
                'variety': selected_v,
                'variety_en': selected_v.split('(')[0].strip(),
                'desc': v_info['desc'],
                'desc_en': v_info['desc_en'],
                'exchange_rate': f"{exchange_rate:,.1f}",
                'unit_price': f"{price:,.2f}",
                'quantity_ton': f"{qty:,.1f}",
                'total_usd': f"{total_usd:,.2f}",
                'total_krw': f"{int(total_krw):,}",
                'weather_en': weather_str,
                'ai_opinion': ""
            }
           
            # AI 분석 실행 (선택된 언어 코드를 넘김)
            ai_advice = get_ai_advice(prop_data, lang_code)
            prop_data['ai_opinion'] = ai_advice
           
            # 상태 저장
            st.session_state['generated_proposal'] = prop_data
            st.session_state['generated_lang'] = lang_code
           
            st.success("제안서 생성 완료!")
            st.rerun()


    # ===========================================
    # 결과물 미리보기 및 다운로드 (문서 형식으로 개선)
    # ===========================================
    if st.session_state['generated_proposal']:
        st.divider()
        st.markdown("### 제안서 미리보기")
       
        prop_data = st.session_state['generated_proposal']
        current_lang = st.session_state['generated_lang']
        is_ko = (current_lang == 'ko')
       
        # 언어별 레이블
        if is_ko:
            title_text = "수입 의사결정 제안서"
            section1_title = "1. 수입 개요"
            section2_title = "2. 비용 및 규모"
            section3_title = "3. 종합 의견 (AI Analysis)"
            label_country = "수입 대상국"
            label_variety = "선택 품종"
            label_rate = "적용 환율"
            label_price = "단가"
            label_volume = "규모"
            label_total = "총액"
            country_display = f"{prop_data['country']} ({prop_data['port']}항)"
            variety_display = prop_data['variety']
        else:
            title_text = "Coffee Import Proposal"
            section1_title = "1. Import Overview"
            section2_title = "2. Cost & Volume"
            section3_title = "3. Recommendations"
            label_country = "Origin Country"
            label_variety = "Coffee Variety"
            label_rate = "Exchange Rate"
            label_price = "Unit Price"
            label_volume = "Volume"
            label_total = "Total"
            country_display = f"{prop_data['country_en']} ({prop_data['port_en']})"
            variety_display = prop_data['variety_en']
       
        # HTML 렌더링
        st.markdown(f"""
<div style="max-width: 900px; margin: 0 auto; background: white; padding: 50px; border: 1px solid #ccc; box-shadow: 0 4px 15px rgba(0,0,0,0.1); border-radius: 8px; font-family: 'Segoe UI', Arial, sans-serif;">
<h1 style="text-align: center; color: #2E5C8A; font-size: 28px; margin-bottom: 10px; border-bottom: 3px solid #2E5C8A; padding-bottom: 15px;">{title_text}</h1>
<p style="text-align: center; color: #666; font-size: 14px; margin-bottom: 40px;">제안일자: {prop_data['date']}</p>
<h2 style="color: #2E5C8A; font-size: 18px; margin-top: 30px; margin-bottom: 15px; border-bottom: 2px solid #E7F0F9; padding-bottom: 8px;">{section1_title}</h2>
<table style="width: 100%; border-collapse: collapse; margin-bottom: 30px; font-size: 14px;">
<tr style="background-color: #E7F0F9;">
<td style="padding: 12px; border: 1px solid #ccc; font-weight: bold; width: 35%;">{label_country}</td>
<td style="padding: 12px; border: 1px solid #ccc; background-color: white;">{country_display}</td>
</tr>
<tr style="background-color: #E7F0F9;">
<td style="padding: 12px; border: 1px solid #ccc; font-weight: bold;">{label_variety}</td>
<td style="padding: 12px; border: 1px solid #ccc; background-color: white;">{variety_display}</td>
</tr>
<tr style="background-color: #E7F0F9;">
<td style="padding: 12px; border: 1px solid #ccc; font-weight: bold;">{label_rate}</td>
<td style="padding: 12px; border: 1px solid #ccc; background-color: white;">{prop_data['exchange_rate']} KRW/USD</td>
</tr>
</table>
<h2 style="color: #2E5C8A; font-size: 18px; margin-top: 30px; margin-bottom: 15px; border-bottom: 2px solid #E7F0F9; padding-bottom: 8px;">{section2_title}</h2>
<ul style="list-style: none; padding: 0; margin-bottom: 30px; font-size: 14px; line-height: 2;">
<li><strong>• {label_price}:</strong> ${prop_data['unit_price']}/kg</li>
<li><strong>• {label_volume}:</strong> {prop_data['quantity_ton']} ton</li>
<li style="color: #C00000; font-weight: bold;"><strong>• {label_total}:</strong> ${prop_data['total_usd']} ({prop_data['total_krw']} KRW)</li>
</ul>
<h2 style="color: #2E5C8A; font-size: 18px; margin-top: 30px; margin-bottom: 15px; border-bottom: 2px solid #E7F0F9; padding-bottom: 8px;">{section3_title}</h2>
<div style="background-color: #f9f9f9; padding: 20px; border-left: 4px solid #2E5C8A; border-radius: 4px; font-size: 14px; line-height: 1.8; color: #333;">
{prop_data['ai_opinion'].replace('1. ', '<p style="margin-bottom: 15px;"><strong>1. </strong>').replace('2. ', '</p><p style="margin-bottom: 15px;"><strong>2. </strong>').replace('3. ', '</p><p style="margin-bottom: 0;"><strong>3. </strong>') + '</p>'}
</div>
</div>
""", unsafe_allow_html=True)
       
        st.divider()
       
        st.markdown("### 결과 다운로드")
        col_d1, col_d2 = st.columns(2)
       
        # 다운로드 시에도 현재 생성된 언어(current_lang)를 전달
        with col_d1:
            pdf_file = create_pdf_proposal(prop_data, lang=current_lang)
            st.download_button(
                "PDF 파일 다운로드 →",
                pdf_file,
                f"Proposal_{prop_data['country_en']}_{prop_data['date']}.pdf",
                "application/pdf",
                use_container_width=True,
                key="dl_pdf"
            )
       
        with col_d2:
            excel_file = create_excel_proposal(prop_data, lang=current_lang)
            st.download_button(
                "엑셀 파일 다운로드 →",
                excel_file,
                f"Proposal_{prop_data['country_en']}_{prop_data['date']}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="dl_excel"
            )


if __name__ == "__main__":
    show()

