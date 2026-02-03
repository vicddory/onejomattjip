# -*- coding: utf-8 -*-
"""
Tab 2: Coffee Beans Analysis - 원두 품종별 분석 및 제안서 생성
"""

import streamlit as st
import pandas as pd
import requests
import os
from datetime import datetime
from dotenv import load_dotenv
from io import BytesIO

# ReportLab imports for PDF
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Openpyxl imports for Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 환경 변수 로드
load_dotenv()

# ==========================================
# 전역 설정
# ==========================================
KOREAN_FONT = 'Helvetica'
USE_KOREAN_FONT = False

def register_korean_font():
    """시스템에 있는 한글 폰트를 찾아서 등록"""
    global KOREAN_FONT, USE_KOREAN_FONT
    
    font_candidates = [
        '/usr/share/fonts/truetype/nanum/NanumGothic.ttf',
        '/usr/share/fonts/truetype/nanum/NanumBarunGothic.ttf',
        'C:/Windows/Fonts/malgun.ttf',
        'C:/Windows/Fonts/gulim.ttf',
        '/System/Library/Fonts/AppleSDGothicNeo.ttc',
        '/Library/Fonts/NanumGothic.ttf'
    ]
    
    for path in font_candidates:
        if os.path.exists(path):
            try:
                font_name = 'KoreanFont'
                pdfmetrics.registerFont(TTFont(font_name, path))
                KOREAN_FONT = font_name
                USE_KOREAN_FONT = True
                return
            except:
                continue

register_korean_font()

# ==========================================
# API 함수
# ==========================================
def get_current_exchange_rate(api_key):
    try:
        if not api_key:
            return None, "API 키 미설정"
        url = f"https://v6.exchangerate-api.com/v6/{api_key}/latest/USD"
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            return response.json()['conversion_rates']['KRW'], "환율 조회 성공"
        return None, "서버 응답 오류"
    except Exception as e:
        return None, str(e)

def get_country_weather(city_name, api_key):
    try:
        if not api_key:
            return {'temp': 0, 'desc_ko': "API키 없음", 'desc_en': "No API Key"}
        url = f"http://api.openweathermap.org/data/2.5/weather?q={city_name}&appid={api_key}&units=metric&lang=en"
        res = requests.get(url, timeout=10).json()
        
        if res.get('cod') != 200:
            return {'temp': 0, 'desc_ko': "정보 없음", 'desc_en': "No Info"}
        
        desc_en = res['weather'][0]['description']
        temp = res['main']['temp']
        
        weather_map = {
            'clear sky': '맑음', 'few clouds': '구름 조금', 'scattered clouds': '구름 낌',
            'broken clouds': '구름 많음', 'overcast clouds': '흐림', 'light rain': '약한 비',
            'moderate rain': '비', 'heavy intensity rain': '강한 비', 'thunderstorm': '뇌우',
            'snow': '눈', 'mist': '안개', 'haze': '연무'
        }
        desc_ko = weather_map.get(desc_en, desc_en)
        
        return {'temp': temp, 'desc_ko': desc_ko, 'desc_en': desc_en}
    except:
        return {'temp': 0, 'desc_ko': "수신 불가", 'desc_en': "Error"}

def get_ai_advice(context_data, lang_code):
    """OpenAI API를 사용하여 전문가 조언 생성"""
    openai_api_key = os.getenv("OPENAI_API_KEY")
    if not openai_api_key:
        return "⚠️ .env 파일에 OPENAI_API_KEY가 설정되지 않았습니다."

    try:
        from openai import OpenAI
        client = OpenAI(api_key=openai_api_key)
        
        target_lang = "Korean" if lang_code == 'ko' else "English"
        
        system_prompt = f"""
        You are a Global Coffee Trade Expert. 
        Analyze the provided data and provide a professional purchasing recommendation in {target_lang}.
        Keep the response concise (within 3-4 sentences) and professional.
        """
        
        user_prompt = f"""
        [Trade Data]
        - Origin: {context_data['country_en']} ({context_data['port_en']})
        - Variety: {context_data['variety_en']}
        - Current Weather: {context_data['weather_en']}
        - Exchange Rate: {context_data['exchange_rate']} KRW/USD
        - Total Volume: {context_data['quantity_ton']} ton
        - Unit Price: ${context_data['unit_price']}/kg
        
        Based on this, write a 'Recommendation' section for a formal proposal.
        """

        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.7
        )
        return response.choices[0].message.content
    except Exception as e:
        return f"AI 분석 중 오류 발생: {str(e)}"

# ==========================================
# 데이터
# ==========================================
def get_coffee_data():
    return {
        "Ethiopia (에티오피아)": {
            "port": "Djibouti", "port_en": "Djibouti Port", "country_en": "Ethiopia",
            "varieties": {
                "Gesha (게이샤)": {"price": 12.5, "desc": "자스민 향과 독보적인 산미", "desc_en": "Jasmine aroma and distinctive acidity"},
                "Yirgacheffe (예가체프)": {"price": 6.8, "desc": "꽃향기와 밝은 산미", "desc_en": "Floral aroma and bright acidity"},
                "Sidamo (시다모)": {"price": 5.8, "desc": "풍부한 과일 향과 바디감", "desc_en": "Fruity aroma with full body"}
            }
        },
        "Kenya (케냐)": {
            "port": "Mombasa", "port_en": "Mombasa Port", "country_en": "Kenya",
            "varieties": {
                "SL28": {"price": 8.5, "desc": "와인 풍미와 강렬한 산미", "desc_en": "Wine-like flavor with intense acidity"},
                "SL34": {"price": 8.2, "desc": "묵직한 바디감과 깊은 풍미", "desc_en": "Full body with deep flavor"},
                "Ruiru 11": {"price": 5.5, "desc": "깔끔한 맛과 적절한 산미", "desc_en": "Balanced acidity with clean taste"}
            }
        },
        "Colombia (콜롬비아)": {
            "port": "Buenaventura", "port_en": "Buenaventura Port", "country_en": "Colombia",
            "varieties": {
                "Typica (티피카)": {"price": 6.5, "desc": "깔끔한 향미와 단맛", "desc_en": "Clean flavor with sweet finish"},
                "Caturra (카투라)": {"price": 5.2, "desc": "풍부한 산미와 중간 바디감", "desc_en": "Rich acidity with medium body"},
                "Castillo (카스티요)": {"price": 5.0, "desc": "부드럽고 베리류 향미", "desc_en": "Smooth with berry notes"}
            }
        },
        "Guatemala (과테말라)": {
            "port": "Puerto Barrios", "port_en": "Puerto Barrios Port", "country_en": "Guatemala",
            "varieties": {
                "Pacamara (파카마라)": {"price": 7.2, "desc": "복합적인 꽃향기와 묵직한 바디", "desc_en": "Complex floral aroma and full body"},
                "Antigua (안티구아)": {"price": 5.4, "desc": "스모키한 향과 초콜릿 풍미", "desc_en": "Smoky aroma with chocolate flavor"},
                "Bourbon (버번)": {"price": 5.2, "desc": "고소함과 산미의 조화", "desc_en": "Nutty sweetness with smooth acidity"}
            }
        },
        "Brazil (브라질)": {
            "port": "Santos", "port_en": "Santos Port", "country_en": "Brazil",
            "varieties": {
                "Bourbon (버번)": {"price": 5.2, "desc": "뛰어난 단맛과 밸런스", "desc_en": "Excellent sweetness and balance"},
                "Catuai (카투아이)": {"price": 4.5, "desc": "가벼운 바디감과 깔끔함", "desc_en": "Light body and clean taste"},
                "Mundo Novo (문도노보)": {"price": 4.2, "desc": "생산성 좋고 밸런스 잡힘", "desc_en": "Productive and well-balanced"}
            }
        },
        "Indonesia (인도네시아)": {
            "port": "Jakarta", "port_en": "Jakarta Port", "country_en": "Indonesia",
            "varieties": {
                "Mandheling (만델링)": {"price": 5.5, "desc": "흙내음과 초콜릿, 묵직함", "desc_en": "Earthy, chocolate notes with heavy body"},
                "Lintong (린통)": {"price": 5.2, "desc": "허브 향과 묵직한 질감", "desc_en": "Herbal aroma with heavy texture"},
                "Gayo (가요)": {"price": 4.9, "desc": "산미와 단맛의 좋은 균형", "desc_en": "Balanced acidity and sweetness"}
            }
        },
        "Vietnam (베트남)": {
            "port": "Ho Chi Minh", "port_en": "Ho Chi Minh Port", "country_en": "Vietnam",
            "varieties": {
                "Excelsa (엑셀사)": {"price": 4.2, "desc": "독특한 과일 향과 타르트 산미", "desc_en": "Unique fruity aroma with tart acidity"},
                "Catimor (카티모르)": {"price": 3.8, "desc": "산미와 쓴맛의 밸런스", "desc_en": "Balanced acidity and bitterness"},
                "Robusta (로부스타)": {"price": 3.2, "desc": "강한 바디감과 구수한 맛", "desc_en": "Strong body with savory taste"}
            }
        },
        "Costa Rica (코스타리카)": {
            "port": "Limon", "port_en": "Limon Port", "country_en": "Costa Rica",
            "varieties": {
                "Villa Sarchi (빌라 사치)": {"price": 7.5, "desc": "우아한 산미와 꽃향기", "desc_en": "Elegant acidity with floral notes"},
                "Caturra (카투라)": {"price": 5.9, "desc": "밝은 산미와 깨끗한 맛", "desc_en": "Bright acidity with clean finish"},
                "Venecia (베네치아)": {"price": 6.2, "desc": "깊은 단맛과 바디감", "desc_en": "Deep sweetness and full body"}
            }
        },
        "Peru (페루)": {
            "port": "Callao", "port_en": "Callao Port", "country_en": "Peru",
            "varieties": {
                "Typica (티피카)": {"price": 5.1, "desc": "은은한 단맛과 깔끔함", "desc_en": "Subtle sweetness with clean finish"},
                "Bourbon (버번)": {"price": 4.8, "desc": "깊은 풍미와 밸런스", "desc_en": "Deep flavor with excellent balance"},
                "Pache (파체)": {"price": 4.5, "desc": "부드럽고 편안한 맛", "desc_en": "Smooth and mild flavor"}
            }
        },
        "Honduras (온두라스)": {
            "port": "Puerto Cortes", "port_en": "Puerto Cortes Port", "country_en": "Honduras",
            "varieties": {
                "Parainema (파라이네마)": {"price": 4.8, "desc": "열대 과일 향과 부드러움", "desc_en": "Tropical fruit aroma and smooth"},
                "Lempira (렘피라)": {"price": 4.1, "desc": "카라멜 단맛과 견과류", "desc_en": "Caramel sweetness with nutty flavor"},
                "Ihcatefe (이카페)": {"price": 3.9, "desc": "밝은 산미와 청량함", "desc_en": "Bright acidity with refreshing finish"}
            }
        }
    }

# ==========================================
# PDF/Excel 생성 함수
# ==========================================
def create_pdf_proposal(data, lang='ko'):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    styles = getSampleStyleSheet()
    font_name = KOREAN_FONT if USE_KOREAN_FONT else 'Helvetica'
    font_name_bold = KOREAN_FONT if USE_KOREAN_FONT else 'Helvetica-Bold'
    
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontName=font_name_bold, fontSize=24, textColor=colors.HexColor('#1F4788'), spaceAfter=20)
    h1_style = ParagraphStyle('H1', parent=styles['Heading1'], fontName=font_name_bold, fontSize=16, textColor=colors.HexColor('#2E5C8A'), spaceAfter=12)
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=font_name, fontSize=11, leading=16)
    
    story = []
    is_ko = (lang == 'ko')
    
    txt = {
        'title': "수입 의사결정 제안서" if is_ko else "Coffee Import Proposal",
        'date': f"제안일자: {data['date']}" if is_ko else f"Date: {data['date']}",
        's1': "1. 수입 개요" if is_ko else "1. Import Overview",
        's2': "2. 비용 및 규모" if is_ko else "2. Cost & Volume",
        's3': "3. 종합 의견 (AI Analysis)" if is_ko else "3. Recommendations (AI Analysis)",
        'footer': "본 제안서는 Coffee Trade Dashboard에서 생성되었습니다." if is_ko else "Generated by Coffee Trade Dashboard."
    }

    story.append(Paragraph(txt['title'], title_style))
    story.append(Paragraph(txt['date'], ParagraphStyle('Date', parent=normal_style, alignment=TA_CENTER, textColor=colors.grey)))
    story.append(Spacer(1, 0.5*cm))
    story.append(Spacer(1, 1*cm))
    
    story.append(Paragraph(txt['s1'], h1_style))
    country_val = f"{data['country']} ({data['port']}항)" if is_ko else f"{data['country_en']} ({data['port_en']})"
    variety_val = data['variety'] if is_ko else data['variety_en']
    
    tbl_data = [
        ["수입 대상국" if is_ko else "Origin", country_val],
        ["선택 품종" if is_ko else "Variety", variety_val],
        ["적용 환율" if is_ko else "Exchange Rate", f"{data['exchange_rate']} KRW/USD"]
    ]
    
    t = Table(tbl_data, colWidths=[4.5*cm, 12.5*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#E7F0F9')),
        ('FONTNAME', (0,0), (-1,-1), font_name),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 1*cm))
    
    story.append(Paragraph(txt['s2'], h1_style))
    story.append(Paragraph(f"• 단가: ${data['unit_price']}/kg", normal_style))
    story.append(Paragraph(f"• 규모: {data['quantity_ton']} ton", normal_style))
    story.append(Paragraph(f"• 총액: ${data['total_usd']} (≈ {data['total_krw']} KRW)", normal_style))
    story.append(Spacer(1, 1*cm))
    
    story.append(Paragraph(txt['s3'], h1_style))
    story.append(Paragraph(data['ai_opinion'], normal_style))
    story.append(Spacer(1, 2*cm))
    story.append(Paragraph(txt['footer'], ParagraphStyle('Footer', parent=normal_style, alignment=TA_CENTER, fontSize=9, textColor=colors.grey)))

    doc.build(story)
    buffer.seek(0)
    return buffer

def create_excel_proposal(data, lang='ko'):
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Proposal"
    
    is_ko = (lang == 'ko')
    title_font = Font(name='맑은 고딕', size=20, bold=True, color='1F4788')
    header_fill = PatternFill(start_color='E7F0F9', end_color='E7F0F9', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    row = 1
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "수입 의사결정 제안서" if is_ko else "Coffee Import Proposal"
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 2
    
    ws[f'A{row}'] = f"Date: {data['date']}"
    row += 2
    
    labels = ["수입 대상국", "선택 품종", "적용 환율"] if is_ko else ["Origin", "Variety", "Exchange Rate"]
    vals = [
        f"{data['country']} ({data['port']}항)" if is_ko else f"{data['country_en']} ({data['port_en']})",
        data['variety'] if is_ko else data['variety_en'],
        f"{data['exchange_rate']} KRW/USD"
    ]
    
    for i, label in enumerate(labels):
        ws[f'A{row}'] = label
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'] = vals[i]
        ws[f'B{row}'].border = thin_border
        row += 1
    
    row += 1
    ws[f'A{row}'] = "단가" if is_ko else "Unit Price"
    ws[f'B{row}'] = f"${data['unit_price']}/kg"
    row += 1
    ws[f'A{row}'] = "규모" if is_ko else "Volume"
    ws[f'B{row}'] = f"{data['quantity_ton']} ton"
    row += 1
    ws[f'A{row}'] = "총액" if is_ko else "Total"
    ws[f'B{row}'] = f"${data['total_usd']} ({data['total_krw']} KRW)"
    ws[f'B{row}'].font = Font(bold=True, color='C00000')
    row += 2
    
    ws[f'A{row}'] = "AI 분석" if is_ko else "AI Analysis"
    row += 1
    ws.merge_cells(f'A{row}:B{row+2}')
    ws[f'A{row}'] = data['ai_opinion']
    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# ==========================================
# 메인 show 함수
# ==========================================
def show():
    """Bean Analysis 페이지를 렌더링하는 메인 함수"""
    
    st.markdown("""<style> 
        div[data-testid="stMetric"] { 
            background-color: #f8f9fa; border-radius: 10px; padding: 15px; 
            box-shadow: 0 2px 4px rgba(0,0,0,0.1); 
        } 
    </style>""", unsafe_allow_html=True)
    
    # 환경 변수
    ex_rate_key = os.getenv("EXCHANGE_RATE")
    weather_key = os.getenv("WEATHER_API_KEY")
    
    data = get_coffee_data()
    
    # 사이드바 - 환율 설정 및 국가 선택
    with st.sidebar:
        st.header("💰 환율 설정")
        
        if 'exchange_source_tab2' not in st.session_state:
            st.session_state['exchange_source_tab2'] = 'manual'

        tab1, tab2 = st.tabs(["📡 실시간 API", "✍️ 수동 입력"])
        with tab1:
            if st.button("환율 가져오기 🔄", key="tab2_rate_btn"):
                rate, msg = get_current_exchange_rate(ex_rate_key)
                if rate:
                    st.session_state['api_rate_tab2'] = rate
                    st.session_state['exchange_source_tab2'] = 'api'
                    st.success(msg)
                else:
                    st.error(msg)
            if 'api_rate_tab2' in st.session_state:
                st.metric("API 환율", f"{st.session_state['api_rate_tab2']:,.2f} 원")

        with tab2:
            manual_rate = st.number_input("직접 입력", value=1450.0, format="%.2f", 
                                          on_change=lambda: st.session_state.update({'exchange_source_tab2': 'manual'}),
                                          key="tab2_manual_rate")

        exchange_rate = st.session_state.get('api_rate_tab2', manual_rate) if st.session_state.get('exchange_source_tab2') == 'api' else manual_rate
        
        st.divider()
        st.markdown(f"## 적용 환율: **{exchange_rate:,.2f} 원**")
        st.divider()
        
        sorted_countries = sorted(list(data.keys()))
        selected_country = st.radio("국가 선택", sorted_countries, label_visibility="collapsed", key="tab2_country")

    # 메인 화면
    st.title(f"☕ Coffee Trade Dashboard: {selected_country.split('(')[0]}")
    st.divider()
    
    country_info = data[selected_country]
    weather_data = get_country_weather(country_info['port'], weather_key)
    
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.subheader(f"📍 산지 정보 ({selected_country})")
        st.info(f"항구: {country_info['port']} / 기상: {weather_data['temp']}°C, {weather_data['desc_ko']}")
        st.markdown("### 🌿 품종 선택")
        selected_v = st.radio("품종", list(country_info['varieties'].keys()), key="tab2_variety")
        v_info = country_info['varieties'][selected_v]
        st.success(f"특징: {v_info['desc']}")
    
    with col2:
        st.subheader("💰 비용 계산기")
        qty = st.number_input("수입 물량 (Ton)", 1.0, 100.0, 1.0, 0.1, key="tab2_qty")
        
        price = v_info['price']
        total_usd = qty * 1000 * price
        total_krw = total_usd * exchange_rate
        
        c1, c2 = st.columns(2)
        c1.metric("단가 (USD/kg)", f"${price:,.2f}")
        c1.metric("총액 (USD)", f"${total_usd:,.2f}")
        c2.metric("중량 (kg)", f"{qty*1000:,.0f} kg")
        c2.metric("총액 (KRW)", f"{int(total_krw):,} 원")
    
    st.divider()
    st.markdown("### 📄 제안서 생성 및 미리보기")
    
    lang_choice = st.radio("문서 언어", ["한국어 (Korean)", "English"], horizontal=True, key="tab2_lang")
    lang_code = 'ko' if "한국어" in lang_choice else 'en'
    
    weather_str_ko = f"{weather_data['temp']}°C, {weather_data['desc_ko']}"
    weather_str_en = f"{weather_data['temp']}°C, {weather_data['desc_en']}"
    
    prop_data = {
        'date': datetime.now().strftime('%Y-%m-%d'),
        'country': selected_country, 'country_en': country_info['country_en'],
        'port': country_info['port'], 'port_en': country_info['port_en'],
        'variety': selected_v, 'variety_en': selected_v.split('(')[0],
        'description': v_info['desc'], 'description_en': v_info['desc_en'],
        'exchange_rate': f"{exchange_rate:,.1f}",
        'unit_price': f"{price:,.2f}",
        'quantity_ton': f"{qty:,.1f}", 'quantity_kg': f"{qty*1000:,.0f}",
        'total_usd': f"{total_usd:,.2f}", 'total_krw': f"{int(total_krw):,}",
        'weather_ko': weather_str_ko,
        'weather_en': weather_str_en,
        'ai_opinion': ""
    }
    
    default_opinion_ko = f"본 제안서는 실시간 국제 시세와 환율을 기반으로 작성되었습니다.\n현지 날씨({weather_str_ko})를 고려할 때, 적정 재고 확보를 위한 신속한 의사결정이 권장됩니다."
    default_opinion_en = f"This proposal is based on real-time data.\nConsidering local weather ({weather_str_en}), prompt decision-making is recommended."
    
    col_ai, col_preview = st.columns([1, 2])
    
    with col_ai:
        st.markdown("#### 🤖 AI 전문가 분석")
        st.info("OpenAI가 현재 상황을 분석하여 전문적인 조언을 생성합니다.")
        
        if 'generated_advice_tab2' not in st.session_state:
            st.session_state['generated_advice_tab2'] = None
        
        if st.button("✨ AI 전문가 자문 받기", use_container_width=True, key="tab2_ai_btn"):
            with st.spinner("전문가가 분석 중입니다..."):
                advice = get_ai_advice(prop_data, lang_code)
                st.session_state['generated_advice_tab2'] = advice
                st.success("분석 완료!")
        
        if st.session_state['generated_advice_tab2']:
            prop_data['ai_opinion'] = st.session_state['generated_advice_tab2']
        else:
            prop_data['ai_opinion'] = default_opinion_ko if lang_code == 'ko' else default_opinion_en
    
    with col_preview:
        prev_title = "수입 의사결정 제안서" if lang_code == 'ko' else "Coffee Import Proposal"
        
        html_preview = f"""
        <div style="padding:20px; border:1px solid #ddd; border-radius:10px; background-color:white; color:black;">
            <h3 style="text-align:center; color:#1F4788; margin-top:0;">{prev_title}</h3>
            <p style="text-align:center; color:grey; font-size:0.9em;">Date: {prop_data['date']}</p>
            <hr style="margin: 10px 0;">
            <p><b>3. Recommendations (AI Analysis)</b></p>
            <div style="background-color:#f9f9f9; padding:15px; border-left: 5px solid #1F4788; font-style: italic;">
                {prop_data['ai_opinion']}
            </div>
        </div>
        """
        st.markdown(html_preview, unsafe_allow_html=True)
    
    st.divider()
    d_col1, d_col2 = st.columns(2)
    file_prefix = f"Proposal_{country_info['country_en']}_{lang_code}"
    
    with d_col1:
        if st.button("PDF 다운로드 📥", use_container_width=True, key="tab2_pdf_btn"):
            pdf_file = create_pdf_proposal(prop_data, lang=lang_code)
            st.download_button("Click to Save PDF", pdf_file, f"{file_prefix}.pdf", "application/pdf", key="tab2_pdf_dl")
    
    with d_col2:
        if st.button("Excel 다운로드 📥", use_container_width=True, key="tab2_excel_btn"):
            excel_file = create_excel_proposal(prop_data, lang=lang_code)
            st.download_button("Click to Save Excel", excel_file, f"{file_prefix}.xlsx", 
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="tab2_excel_dl")
